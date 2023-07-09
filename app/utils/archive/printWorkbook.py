import openpyxl
import json

def workbook_to_json(workbook):
    data = {}
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        sheet_data = []
        for row in sheet.iter_rows(values_only=True):
            row_data = [str(cell) for cell in row]
            sheet_data.append(row_data)
        data[sheet_name] = sheet_data
    return json.dumps(data, indent=4)

def main():
    file_path = '/home/jgracias/Snapshot/tempDB/spreadsheet/may2023/Jobs_Report.xlsx'
    workbook = openpyxl.load_workbook(file_path)
    json_data = workbook_to_json(workbook)
    print(json_data)

if __name__ == '__main__':
    main()
