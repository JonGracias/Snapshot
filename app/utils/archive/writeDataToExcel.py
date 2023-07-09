import openpyxl

def write_data_to_excel(data):
    # Create a new workbook
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Write column headers
    headers = list(data["RadGridExport-53"][0].keys())
    for col_idx, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=col_idx).value = header

    # Write data rows
    for row_idx, entry in enumerate(data["RadGridExport-53"], start=2):
        for col_idx, header in enumerate(headers, start=1):
            sheet.cell(row=row_idx, column=col_idx).value = entry[header]

    workbook.save('new_spreadsheet.xlsx')

    # Print the content of the worksheet
    """ for row in sheet.iter_rows(values_only=True):
        print(row) """

    return workbook
