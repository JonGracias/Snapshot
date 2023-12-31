from flask import Flask, render_template, request,redirect,jsonify
from collections import defaultdict 
from flask_pymongo import PyMongo
from operator import itemgetter
from flask_caching import Cache
from datetime import datetime
from flask_cors import CORS
from flask import Markup
import pandas as pd
import openpyxl
import dotenv
import flask
import json
import os
import re

dotenv.load_dotenv()
app = Flask(__name__)
CORS(app)
app.config.update(
    TESTING=True,
    TEMPLATES_AUTO_RELOAD=True,
    DEBUG=True,
    MONGO_URI=os.getenv("DATABASE_URL"),
    SECRET_KEY=os.getenv("SECRET_KEY"),
    CACHE_TYPE = 'simple',
)


mongo = PyMongo(app)
cache = Cache(app)
current_user = "user"  
collection_name = f"{current_user}" 

#--------------------------------------Routes-----------------------------------------------------#
@app.route('/save', methods=['GET'])
def save():
    return render_template('save.html')

# For React
@app.route('/navigation', methods=['POST'])
def navigation():
    clicked_items = request.json.get('clickedItems', [])
    titles, dates, tables = Spreadsheet.get(clicked_items)
    return jsonify(
        titles=titles,
        dates=dates,
        tables=tables,
    )

# For React Front End
@app.route('/browseDB', methods=['GET'])
def browse():
    query = ''
    results_html = format_for_browse(query)
    response_data = results_html
    return jsonify(response_data)


@app.route('/search', methods=['POST'])
def search_results():
    query = request.form['query']
    results_html = format_for_browse(query)
    return render_template('load.html', results_html=results_html)

@app.route('/upload', methods=['POST'])
def upload():
    file = request.files['file']
    date = request.form['dateInput']

    workbook = openpyxl.load_workbook(file, data_only=True)
    json_data = format_workbook(workbook)

    for sheet_name, sheet_data in json_data.items():
        position = generate_position_number(sheet_name)
        title = sheet_name
        formatted_data = json.dumps({sheet_name: sheet_data}, indent=4)
        Spreadsheet.save(formatted_data, title, date, position)
    return flask.redirect(flask.url_for('index'))

@app.route('/', defaults={'path': ''})
@app.route('/<path:path>')
def index(path):
    return render_template('index.html')


#---------------------------------Functions and Classes----------------------#
class Spreadsheet:
    @staticmethod
    def save(table, title, date, position):
        id = f"{title}-{date}"
        timestamp = datetime.now()
        filter_query = {"_id":id}
        existing_document = mongo.db[collection_name].find_one(filter_query)

        if existing_document:
            update_operation = {"$set": {"table": table}}
            mongo.db[collection_name].update_one(filter_query, update_operation)
        else:
            data = {"_id":id, "title": title, "position": position, "date":date, "timestamp": timestamp, "table": table}
            mongo.db[collection_name].insert_one(data)





    @staticmethod
    def get(clicked_items=None):
        if clicked_items:
            """ print('Clicked Items Before: ' + str(clicked_items)) """
            # Retrieve MongoDB documents based on clicked_items
            documents = list(
                mongo.db[collection_name].find({'title': {'$in': clicked_items}})
            )
        else:
            documents = list(mongo.db[collection_name].find())

        title_counts = -1
        isNewTitle = []
        titles = []
        grouped_dates = []
        dates = []  
        grouped_tables = []
        tables = []  
       


        # Function to get the index of a title in the clicked_items list
        def get_clicked_items_index(title):
            return clicked_items.index(title) if title in clicked_items else -1


        # Custom sorting function to sort documents based on clicked_items order or by date
        def custom_sort(document):
            return get_clicked_items_index(document['title']), document['date']

        # Sort documents based on clicked_items order or by date
        documents.sort(key=custom_sort)

        for document in documents:
            title = document['title']
            date = document['date']
            table = document['table']
            content_dict = json.loads(table)

            content_key = next(iter(content_dict.keys()))
            if isinstance(content_dict[content_key], str):
                transformed_content = [{k: v.replace('\n', '')} for k, v in json.loads(content_dict[content_key]).items()]
            else:
                transformed_content = [
                    {k: v.replace('\n', '') for k, v in entry.items()} for entry in content_dict[content_key]
                ]

            # Convert the list of dictionaries to a pandas DataFrame
            if isinstance(transformed_content, str):
                df = pd.DataFrame()
            else:
                df = pd.DataFrame(transformed_content)
            dataSource = df.to_html(index=False)
        

            # Increment the count only on the first occurrence of the title
            if title not in isNewTitle:
                title_counts += 1
                # Append the title to the titles list
                titles.append(title)
                dates=[]
                grouped_dates.insert(title_counts, dates)
                tables = []
                grouped_tables.insert(title_counts, dataSource)

            # Add the index number to the document dictionary
            document['index'] = title_counts


            # Add current title to isNewTitle
            dates.append(date)
            tables.append(dataSource)
            grouped_tables[title_counts] = tables
            grouped_dates[title_counts]=dates
            isNewTitle.append(title)



        if documents and 'table' in documents[0]:
            print('Tables' + str(grouped_tables))
            return titles, grouped_dates, grouped_tables
        else:
            print("No documents found or missing 'table' field")
            return [
                {
                    '_id': 'None',
                    'title': 'No Data Found',
                    'position': 0,
                    'date': '00/0000',
                    'table': '{\n    "Sheet1": [{\n"Load": "To Get",\n"Data": "Started"\n}]}',
                }
            ]


def format_for_html(clicked_items):
    int_dates = []
    names = []
    if clicked_items:

        data = Spreadsheet.get(clicked_items)

        # Create a defaultdict with list as default value
        int_dict = defaultdict(dict)

        """ for document in data:
            name = document.pop('title')
            table = document.pop('table')
            date = document.pop('date')
            content_dict = json.loads(table)

            content_key = next(iter(content_dict.keys()))
            if isinstance(content_dict[content_key], str):
                transformed_content = [{k: v.replace('\n', '')} for k, v in json.loads(content_dict[content_key]).items()]
            else:
                transformed_content = [
                    {k: v.replace('\n', '') for k, v in entry.items()} for entry in content_dict[content_key]
                ]

            if isinstance(transformed_content, str):
                df = pd.DataFrame()
            else:
                df = pd.DataFrame(transformed_content)

            #document['table'] = df.to_html(index=False)
            document['table'] = 'Newtable'
            document['date'] = date
            document['name'] = name


            # Append the document to the list under its index key in the defaultdict
            int_dict[document['index']][document['date']] = document
            # Extract unique dates from the result dictionary and convert them to a sorted list
            int_dates[document['index']].append(date)
            names[document['index']].append(name)
            
    else:
        int_dict = {0: {'05/2023': {'title': 'Default', 'table': 'Default 1', 'date': '05/2023'}},
                     1: {'06/2023': {'title': 'Default2', 'table': 'Default 2', 'date': '05/2023'}}}
        int_dates = {0: ['05/2023','06/2023'], 1: ['05/2024','06/2024']}
     """
    return int_dates, names



def format_for_browse(query):
    if query:
        regex_query = re.compile(f'.*{query}.*', re.IGNORECASE)
        documents = mongo.db[collection_name].find({'title': regex_query})
    else:
        documents = mongo.db[collection_name].find({}, {"title": 1, "_id": 0})
    
    data = [doc["title"] for doc in documents]
    titles = set(data)
    
    unique_titles = sorted(list(titles))

    if unique_titles:
        job_list = []
        for index, title in enumerate(unique_titles):
            job_name = title
            job_list.append(job_name)
        
        if query:
            results_list = [f'Search Results for "{query}":']
            results_list.extend(job_list)
        else:
            results_list = job_list
    else:
        if query:
            results_list = [f'No results found for "{query}"']
        else:
            results_list = ['Nothing in Database.']
            
    return results_list



def generate_position_number(sheet_name):
    result = mongo.db[collection_name].find_one({'title': sheet_name})

    if result:
        position = result['position']
    else:
        position = largest_position_number()

    return position

def largest_position_number():
    pipeline = [
        {"$group": {"_id": None, "max_number": {"$max": "$position"}}},
        {"$limit": 1}
    ]

    aggregation_result = mongo.db[collection_name].aggregate(pipeline)
    newest_document = next(aggregation_result, None)

    max_number = newest_document['max_number']+1 if newest_document else 0
    return max_number

def format_workbook(workbook):
    data = {}
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        sheet_data = []
        headers = [cell.value for cell in sheet[1]]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_data = {header: str(cell) for header, cell in zip(headers, row)}
            filtered_row_data = {k: v for k, v in row_data.items() if v is not None and k is not None}
            if filtered_row_data:
                sheet_data.append(filtered_row_data)
        data[sheet_name] = sheet_data
    return data

if __name__ == '__main__':
    app.run(host='localhost', port=5000)
